#!/usr/bin/env python3
"""
Import 运镜 Prompt 精品库 Excel → knowledge/camera/*.json

Default source:
  E:\\AI\\知识库\\提示词\\运镜Prompt精品库_清洗版.xlsx

Usage:
  python scripts/import_camera_xlsx.py
  python scripts/import_camera_xlsx.py --xlsx "D:\\path\\to\\file.xlsx"
  python scripts/import_camera_xlsx.py --xlsx "..." --out knowledge/camera
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"E:\AI\知识库\提示词\运镜Prompt精品库_清洗版.xlsx")
DEFAULT_OUT = ROOT / "knowledge" / "camera"

SHEET_MOVES = "运镜Prompt精品库"
SHEET_SIZES = "景别词库_非运镜"

# Map free-text 叙事/情绪 → pipeline emotion keys
EMOTION_KEYWORDS: dict[str, list[str]] = {
    "oppression": [
        "压迫",
        "压制",
        "逼近",
        "窒息",
        "威胁",
        "控制",
        "威压",
        "dominance",
        "pressure",
        "intimidate",
    ],
    "suspicion": [
        "怀疑",
        "猜疑",
        "窥视",
        "监视",
        "试探",
        "不安",
        "紧张",
        "suspicion",
        "unease",
        "paranoi",
        "watch",
    ],
    "intimacy": [
        "亲密",
        "柔情",
        "温情",
        "靠近",
        "私密",
        "告白",
        "intimacy",
        "tender",
        "romantic",
        "close",
    ],
    "grief": [
        "悲伤",
        "哀悼",
        "失落",
        "沉重",
        "绝望",
        "眼泪",
        "grief",
        "sorrow",
        "mourn",
        "melanchol",
    ],
    "revelation": [
        "揭示",
        "发现",
        "真相",
        "揭晓",
        "亮相",
        "信息",
        "reveal",
        "discovery",
        "twist",
        "uncover",
    ],
    "calm": [
        "平静",
        "日常",
        "冷静",
        "舒缓",
        "建立",
        "交代环境",
        "展示空间",
        "calm",
        "peaceful",
        "establish",
        "observ",
    ],
    "dread": [
        "恐惧",
        "惊悚",
        "恐怖",
        "眩晕",
        "失衡",
        "危险",
        "不安",
        "dread",
        "horror",
        "thriller",
        "disorient",
        "tense",
    ],
}


def _slug(en: str, idx: int) -> str:
    base = re.sub(r"[^a-zA-Z0-9]+", "_", (en or "").strip()).strip("_").lower()
    if not base:
        base = f"move_{idx}"
    return base[:64]


def _split_aliases(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[;；,/|、]+", str(raw))
    return [p.strip() for p in parts if p and p.strip()]


def _split_scenes(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[;；,/|、]+", str(raw))
    return [p.strip() for p in parts if p and p.strip()]


def _emotions_for_text(text: str) -> list[str]:
    t = (text or "").lower()
    # keep original for Chinese match
    raw = text or ""
    hit: list[str] = []
    for emo, kws in EMOTION_KEYWORDS.items():
        for kw in kws:
            if kw.lower() in t or kw in raw:
                hit.append(emo)
                break
    return hit


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        import subprocess

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

    return openpyxl


def load_moves(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    # expected columns
    col = {name: i for i, name in enumerate(header)}
    required = ["英文词条", "中文词条", "英文Prompt示例"]
    for r in required:
        if r not in col:
            raise ValueError(f"运镜表缺少列: {r}; 实际: {header}")

    items: list[dict] = []
    for row in rows[1:]:
        if not row or row[col.get("序号", 0)] is None and not row[col["英文词条"]]:
            continue
        en = str(row[col["英文词条"]] or "").strip()
        if not en:
            continue
        idx = row[col["序号"]] if "序号" in col else len(items) + 1
        try:
            idx_i = int(idx)
        except (TypeError, ValueError):
            idx_i = len(items) + 1

        zh = str(row[col["中文词条"]] or "").strip() if "中文词条" in col else ""
        aliases = _split_aliases(row[col["同义词/别名"]] if "同义词/别名" in col else None)
        mtype = str(row[col["运镜类型"]] or "").strip() if "运镜类型" in col else ""
        scenes = _split_scenes(row[col["应用场景"]] if "应用场景" in col else None)
        narrative = (
            str(row[col["叙事/情绪用途"]] or "").strip() if "叙事/情绪用途" in col else ""
        )
        prompt = (
            str(row[col["英文Prompt示例"]] or "").strip() if "英文Prompt示例" in col else ""
        )
        note = str(row[col["中文说明"]] or "").strip() if "中文说明" in col else ""

        emotions = _emotions_for_text(narrative + " " + note + " " + mtype)
        mid = _slug(en, idx_i)

        items.append(
            {
                "id": mid,
                "index": idx_i,
                "en": en,
                "zh": zh,
                "aliases": aliases,
                "move_type": mtype,
                "scenes": scenes,
                "narrative_emotion": narrative,
                "emotions": emotions,
                "prompt_en": prompt,
                "note_zh": note,
            }
        )
    return items


def load_shot_sizes(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = {name: i for i, name in enumerate(header)}
    items: list[dict] = []
    for row in rows[1:]:
        if not row:
            continue
        # skip说明 row
        first = row[0]
        if first is None or (isinstance(first, str) and first.startswith("说明")):
            continue
        try:
            idx_i = int(first)
        except (TypeError, ValueError):
            continue
        en = str(row[col.get("英文", 1)] or "").strip()
        zh = str(row[col.get("中文", 2)] or "").strip()
        aliases = _split_aliases(row[col.get("缩写/别名", 3)] if "缩写/别名" in col else row[3])
        desc = str(row[col.get("说明", 4)] or "").strip() if "说明" in col else ""
        remark = str(row[col.get("备注", 5)] or "").strip() if "备注" in col else ""
        # code: prefer first alias that looks like EWS/WS/MCU
        code = ""
        for a in aliases:
            if re.fullmatch(r"[A-Z]{2,4}", a.replace("-", "")) or re.fullmatch(
                r"[A-Z]{1,4}", a
            ):
                code = a.split()[0] if " " in a else a
                break
        if not code and aliases:
            code = aliases[0]
        items.append(
            {
                "id": _slug(en or code or f"size_{idx_i}", idx_i),
                "index": idx_i,
                "en": en,
                "zh": zh,
                "code": code,
                "aliases": aliases,
                "description": desc,
                "remark": remark,
            }
        )
    return items


def build_emotion_index(moves: list[dict]) -> dict[str, list[str]]:
    idx: dict[str, list[str]] = defaultdict(list)
    for m in moves:
        for emo in m.get("emotions") or []:
            if m["id"] not in idx[emo]:
                idx[emo].append(m["id"])
    # ensure all pipeline keys exist
    for k in EMOTION_KEYWORDS:
        idx.setdefault(k, [])
    return dict(idx)


def build_moves_compact(moves: list[dict]) -> dict[str, dict]:
    """Compat shape for existing timing/move_durations consumers (by movement type key)."""
    compact: dict[str, dict] = {}
    for m in moves:
        key = _slug(m["en"], m["index"])
        # duration hints by type bucket
        mt = m.get("move_type") or ""
        if "基础" in mt:
            min_s, pref = 2.0, 3.5
        elif "复合" in mt:
            min_s, pref = 3.5, 5.5
        elif "光学" in mt:
            min_s, pref = 2.0, 3.0
        elif "特殊" in mt:
            min_s, pref = 2.5, 4.0
        else:
            min_s, pref = 2.5, 4.0
        compact[key] = {
            "en": m["en"],
            "zh": m["zh"],
            "min_sec": min_s,
            "preferred_sec": pref,
            "prompt_hint": m.get("prompt_en") or m["en"],
            "use_when": m.get("emotions") or [],
            "move_type": m.get("move_type"),
        }
    # keep legacy simple keys used by stubs
    legacy_alias = {
        "static": "static_hold",
        "static_hold": "static_hold",
        "slow_push_in": "slow_push_in",
        "very_slow_push_in": "very_slow_push_in",
        "micro_drift": "micro_drift",
        "slow_lateral": "slow_lateral",
        "creep_in": "creep_in",
        "rack_focus_support": "rack_focus_support",
    }
    for k, v in list(compact.items()):
        en_l = (v.get("en") or "").lower()
        if "dolly in" in en_l or en_l == "push in":
            legacy_alias["slow_push_in"] = k
        if "static" in en_l or en_l == "locked off":
            legacy_alias["static_hold"] = k
    compact["_legacy_aliases"] = legacy_alias
    compact["default"] = {
        "min_sec": 2.0,
        "preferred_sec": 3.5,
        "prompt_hint": "controlled camera move",
    }
    return compact


def build_emotion_to_camera(moves: list[dict], emotion_index: dict[str, list[str]]) -> dict:
    """Enrich pipeline emotion_to_camera with catalog ids + sample prompts."""
    by_id = {m["id"]: m for m in moves}
    out: dict = {}
    for emo, ids in emotion_index.items():
        sample = []
        angles = []
        lenses = [35, 40, 50]
        avoid = ["unmotivated_dutch", "orbit"] if emo in {"intimacy", "calm"} else ["beauty_orbit"]
        preferred_moves = []
        for mid in ids[:12]:
            m = by_id.get(mid)
            if not m:
                continue
            preferred_moves.append(m["en"])
            sample.append(
                {
                    "id": mid,
                    "en": m["en"],
                    "zh": m["zh"],
                    "prompt_en": m.get("prompt_en"),
                }
            )
        # angle heuristics
        if emo in {"oppression", "dread"}:
            angles = ["slight_low", "eye_level_tight"]
            lenses = [28, 35, 40]
        elif emo == "intimacy":
            angles = ["eye_level"]
            lenses = [50, 65, 85]
        elif emo == "revelation":
            angles = ["eye_level", "slight_low"]
            lenses = [35, 40, 50]
        else:
            angles = ["eye_level"]
            lenses = [35, 40, 50]

        out[emo] = {
            "preferred_angles": angles,
            "preferred_moves": preferred_moves[:8] or ["static hold", "slow dolly in"],
            "preferred_move_ids": ids[:12],
            "lens_mm": lenses,
            "avoid": avoid,
            "catalog_samples": sample[:6],
        }
    return out


def write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"  wrote {path} ({path.stat().st_size} bytes)")


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Import camera move Excel into knowledge/camera")
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Source xlsx path")
    p.add_argument("--out", type=Path, default=DEFAULT_OUT, help="Output knowledge/camera dir")
    p.add_argument(
        "--merge-emotion",
        action="store_true",
        default=True,
        help="Write emotion_to_camera_from_catalog.json (default on)",
    )
    args = p.parse_args(argv)

    xlsx: Path = args.xlsx
    out: Path = args.out
    if not xlsx.is_absolute():
        xlsx = (ROOT / xlsx).resolve()
    if not out.is_absolute():
        out = (ROOT / out).resolve()

    if not xlsx.exists():
        print(f"[error] xlsx not found: {xlsx}", file=sys.stderr)
        return 1

    openpyxl = _require_openpyxl()
    print(f"Reading {xlsx}")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    if SHEET_MOVES not in wb.sheetnames:
        print(f"[error] missing sheet {SHEET_MOVES}: {wb.sheetnames}", file=sys.stderr)
        return 1

    moves = load_moves(wb[SHEET_MOVES])
    sizes = load_shot_sizes(wb[SHEET_SIZES]) if SHEET_SIZES in wb.sheetnames else []
    wb.close()

    print(f"Parsed moves={len(moves)} shot_sizes={len(sizes)}")
    emotion_index = build_emotion_index(moves)
    for emo, ids in emotion_index.items():
        print(f"  emotion[{emo}]: {len(ids)} moves")

    meta = {
        "source_xlsx": str(xlsx),
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "move_count": len(moves),
        "shot_size_count": len(sizes),
        "sheets": [SHEET_MOVES, SHEET_SIZES],
        "note": "Excel is edit source; runtime should read these JSON files.",
    }

    catalog = {
        "meta": meta,
        "moves": moves,
        "by_type": {},
    }
    by_type: dict[str, list[str]] = defaultdict(list)
    for m in moves:
        by_type[m.get("move_type") or "未分类"].append(m["id"])
    catalog["by_type"] = dict(by_type)

    write_json(out / "moves_catalog.json", catalog)
    write_json(out / "shot_sizes.json", {"meta": meta, "sizes": sizes})
    write_json(out / "emotion_move_index.json", {"meta": meta, "index": emotion_index})

    # Compact moves for duration hints (does not delete hand-tuned move_durations.json)
    write_json(out / "moves_from_catalog.json", build_moves_compact(moves))

    if args.merge_emotion:
        etc = build_emotion_to_camera(moves, emotion_index)
        write_json(out / "emotion_to_camera_from_catalog.json", etc)
        # Also refresh main emotion_to_camera.json used by pipeline (backup-friendly overwrite)
        write_json(out / "emotion_to_camera.json", etc)

    # Human-readable index
    md_lines = [
        f"# Camera catalog (imported)",
        "",
        f"- source: `{xlsx}`",
        f"- moves: **{len(moves)}**",
        f"- shot sizes: **{len(sizes)}**",
        f"- imported: {meta['imported_at']}",
        "",
        "## By type",
        "",
    ]
    for t, ids in sorted(by_type.items(), key=lambda x: -len(x[1])):
        md_lines.append(f"- {t}: {len(ids)}")
    md_lines += ["", "## Emotion index sizes", ""]
    for emo, ids in emotion_index.items():
        md_lines.append(f"- {emo}: {len(ids)}")
    md_lines += ["", "## Sample moves", ""]
    for m in moves[:15]:
        md_lines.append(f"- **{m['en']}** / {m['zh']}: `{m['prompt_en'][:80]}...`" if len(m.get("prompt_en") or "") > 80 else f"- **{m['en']}** / {m['zh']}: `{m.get('prompt_en')}`")
    (out / "CATALOG_README.md").write_text("\n".join(md_lines) + "\n", encoding="utf-8")
    print(f"  wrote {out / 'CATALOG_README.md'}")

    print("Done. Runtime knowledge is under:", out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
